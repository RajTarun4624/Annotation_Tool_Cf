import io
import csv
from datetime import UTC, datetime
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.dependencies import get_current_user, require_permission
from app.schemas.pagination import Page
from app.crud.role import get_role_by_id, get_role_by_name
from app.crud.user import (
    create_user,
    get_user_by_email,
    get_user_by_id,
    list_users,
    update_user,
    update_user_status,
)
from app.schemas.user import (
    UserCreateRequest,
    UserProfileResponse,
    UserStatusUpdateRequest,
    UserSummaryResponse,
    UserUpdateRequest,
)

router = APIRouter()


@router.get("/me", response_model=UserProfileResponse)
def read_my_profile(current_user: Annotated[dict, Depends(get_current_user)]) -> UserProfileResponse:
    profile_payload = {key: value for key, value in current_user.items() if key != "hashed_password"}
    return UserProfileResponse.model_validate(profile_payload)


@router.get("/", response_model=Page[UserSummaryResponse])
def read_users(
    _: Annotated[dict, Depends(require_permission("users"))],
    db: Annotated[Session, Depends(get_db)],
    search: str | None = Query(default=None),
    page: int | None = Query(default=None, ge=1),
    page_size: int | None = Query(default=None, ge=1, le=500),
) -> Page[UserSummaryResponse]:
    users, total, eff_page, eff_size = list_users(db, search=search, page=page, page_size=page_size)
    return Page[UserSummaryResponse](
        items=[
            UserSummaryResponse.model_validate({key: value for key, value in user.items() if key != "hashed_password"})
            for user in users
        ],
        total=total, page=eff_page, page_size=eff_size,
    )


@router.get("/export")
def export_users(
    _: Annotated[dict, Depends(require_permission("users"))],
    db: Annotated[Session, Depends(get_db)],
    search: str | None = Query(default=None),
) -> StreamingResponse:
    """Export the (optionally search-filtered) user list as an .xlsx file.

    Declared BEFORE `/{user_id}` so the literal path isn't captured as a user id.
    """
    users, _total, _page, _size = list_users(db, search=search)

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = ["#", "Full Name", "Email", "Role", "Permissions", "Status", "Created At"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for i, u in enumerate(users, start=1):
        created = u.get("created_at")
        created_str = created.strftime("%Y-%m-%d %H:%M") if hasattr(created, "strftime") else (str(created) if created else "")
        ws.append([
            i,
            u.get("full_name", ""),
            u.get("email", ""),
            u.get("role_name") or "—",
            ", ".join(u.get("permissions") or []),
            "Active" if u.get("is_active") else "Inactive",
            created_str,
        ])

    widths = [6, 26, 34, 20, 44, 12, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"users-{datetime.now(UTC).strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/import-template")
def import_template(
    _: Annotated[dict, Depends(require_permission("users"))],
) -> StreamingResponse:
    """Download an Excel template for bulk user imports."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"
    headers = ["Full Name", "Email", "Password", "Role"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    ws.append(["John Doe", "john.doe@example.com", "Password123", "User"])  # sample row uses a real role (Admin | QA | User)

    widths = [20, 30, 20, 20]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="users-import-template.xlsx"'},
    )


@router.post("/import")
def import_users(
    _: Annotated[dict, Depends(require_permission("users"))],
    db: Annotated[Session, Depends(get_db)],
    file: UploadFile = File(...),
):
    """Import users from a CSV or Excel (.xlsx) file."""
    filename = file.filename or ""
    is_xlsx = filename.endswith(".xlsx")
    is_csv = filename.endswith(".csv")

    if not (is_xlsx or is_csv):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file format. Please upload a .csv or .xlsx file."
        )

    errors = []
    data_rows = []

    try:
        if is_xlsx:
            content = file.file.read()
            buffer = io.BytesIO(content)
            wb = load_workbook(buffer, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Filter out completely empty rows
                if any(cell is not None for cell in row):
                    rows.append(list(row))
            
            if not rows:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file is empty.")
            
            headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
            data_rows = rows[1:]
        else:
            content = file.file.read().decode("utf-8-sig", errors="ignore")
            reader = csv.reader(io.StringIO(content))
            rows = [row for row in reader if any(cell.strip() for cell in row)]
            if not rows:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV file is empty.")
            
            headers = [str(cell).strip().lower() for cell in rows[0]]
            data_rows = rows[1:]
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {str(e)}"
        )

    # Map headers to indices
    col_map = {}
    for idx, h in enumerate(headers):
        if h in ["full name", "fullname", "name"]:
            col_map["full_name"] = idx
        elif h in ["email", "e-mail"]:
            col_map["email"] = idx
        elif h in ["password"]:
            col_map["password"] = idx
        elif h in ["role", "assigned role", "role name", "role_name"]:
            col_map["role"] = idx

    required = ["full_name", "email", "password", "role"]
    missing_cols = [r.replace('_', ' ').title() for r in required if r not in col_map]
    if missing_cols:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required columns: {', '.join(missing_cols)}."
        )

    imported_count = 0
    roles_cache = {}

    for row_idx, row in enumerate(data_rows, start=2):
        try:
            full_name = str(row[col_map["full_name"]]).strip() if col_map["full_name"] < len(row) and row[col_map["full_name"]] is not None else ""
            email = str(row[col_map["email"]]).strip() if col_map["email"] < len(row) and row[col_map["email"]] is not None else ""
            password = str(row[col_map["password"]]).strip() if col_map["password"] < len(row) and row[col_map["password"]] is not None else ""
            role_name = str(row[col_map["role"]]).strip() if col_map["role"] < len(row) and row[col_map["role"]] is not None else ""
        except Exception:
            errors.append(f"Row {row_idx}: Failed to read columns.")
            continue

        if not full_name or not email or not password or not role_name:
            missing_fields = []
            if not full_name: missing_fields.append("Full Name")
            if not email: missing_fields.append("Email")
            if not password: missing_fields.append("Password")
            if not role_name: missing_fields.append("Role")
            errors.append(f"Row {row_idx}: Missing values for: {', '.join(missing_fields)}")
            continue

        if len(password) < 8:
            errors.append(f"Row {row_idx}: Password must be at least 8 characters.")
            continue

        if "@" not in email or "." not in email:
            errors.append(f"Row {row_idx}: Invalid email address format.")
            continue

        role_key = role_name.lower()
        if role_key not in roles_cache:
            role_db = get_role_by_name(db, role_name)
            roles_cache[role_key] = role_db

        role = roles_cache[role_key]
        if not role:
            errors.append(f"Row {row_idx}: Role '{role_name}' does not exist or is inactive.")
            continue

        if get_user_by_email(db, email.lower()):
            errors.append(f"Row {row_idx}: User with email '{email}' already exists.")
            continue

        try:
            create_user(db, {
                "full_name": full_name,
                "email": email,
                "password": password,
                "role_id": role["id"],
                "is_active": True
            })
            imported_count += 1
        except Exception as ex:
            errors.append(f"Row {row_idx}: Internal error creating user: {str(ex)}")

    return {"imported_count": imported_count, "errors": errors}


@router.get("/{user_id}", response_model=UserProfileResponse)
def read_user(
    user_id: str,
    _: Annotated[dict, Depends(require_permission("users"))],
    db: Annotated[Session, Depends(get_db)],
) -> UserProfileResponse:
    user = get_user_by_id(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    return UserProfileResponse.model_validate({key: value for key, value in user.items() if key != "hashed_password"})


@router.post("/", response_model=UserProfileResponse, status_code=status.HTTP_201_CREATED)
def add_user(
    payload: UserCreateRequest,
    _: Annotated[dict, Depends(require_permission("users"))],
    db: Annotated[Session, Depends(get_db)],
) -> UserProfileResponse:
    if get_user_by_email(db, payload.email.lower()):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Email already registered")
    role = get_role_by_id(db, payload.role_id)
    if not role:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assigned role not found")
    created = create_user(db, payload.model_dump())
    return UserProfileResponse.model_validate({key: value for key, value in created.items() if key != "hashed_password"})


@router.put("/{user_id}", response_model=UserProfileResponse)
def edit_user(
    user_id: str,
    payload: UserUpdateRequest,
    _: Annotated[dict, Depends(require_permission("users"))],
    db: Annotated[Session, Depends(get_db)],
) -> UserProfileResponse:
    existing = get_user_by_id(db, user_id)
    if not existing:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    duplicate = get_user_by_email(db, payload.email.lower())
    if duplicate and duplicate["id"] != user_id:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Email already registered")
    role = get_role_by_id(db, payload.role_id)
    if not role:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assigned role not found")
    updated = update_user(db, user_id, payload.model_dump())
    if not updated:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unable to update user")
    return UserProfileResponse.model_validate({key: value for key, value in updated.items() if key != "hashed_password"})


@router.patch("/{user_id}/status", response_model=UserProfileResponse)
def edit_user_status(
    user_id: str,
    payload: UserStatusUpdateRequest,
    _: Annotated[dict, Depends(require_permission("users"))],
    db: Annotated[Session, Depends(get_db)],
) -> UserProfileResponse:
    updated = update_user_status(db, user_id, payload.is_active)
    if not updated:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    return UserProfileResponse.model_validate({key: value for key, value in updated.items() if key != "hashed_password"})
