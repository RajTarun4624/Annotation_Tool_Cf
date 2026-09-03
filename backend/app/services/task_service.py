import io
import re
from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.models.queue import Queue
from app.models.task import Task
from app.repositories.task_repository import TaskRepository

_WHITESPACE_RE = re.compile(r"\s+")


def format_aht(seconds: int | None) -> str:
    if seconds is None or seconds < 0:
        return "0s"
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    parts = []
    if h > 0:
        parts.append(f"{h}h")
    if m > 0 or h > 0:
        parts.append(f"{m}m")
    parts.append(f"{s}s")
    return " ".join(parts)


def format_dt(dt: datetime | None) -> str:
    if not dt:
        return "—"
    # Format: MMM D, YY HH:mm (e.g. Jun 5, 26 17:20)
    return dt.strftime("%b %d, %y %H:%M")


def format_aht_hhmmss(seconds: int | None) -> str:
    if seconds is None or seconds < 0:
        return "00:00:00"
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def format_excel_dt(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def format_iso(dt: datetime | None) -> str | None:
    if not dt:
        return None
    iso = dt.isoformat()
    if not iso.endswith("Z") and "+00:00" not in iso:
        return iso + "Z"
    return iso


def input_preview(text: str | None, limit: int = 140) -> str:
    """First ``limit`` characters of the prompt collapsed onto one line."""
    if not text:
        return ""
    flat = _WHITESPACE_RE.sub(" ", str(text)).strip()
    if len(flat) <= limit:
        return flat
    return flat[:limit].rstrip() + "…"


def serialize_task(task: Task) -> dict:
    # `queue.task_name` / `queue.name` are nullable columns, so coalesce every
    # required string field to a real string — a null here would otherwise fail
    # TaskItemResponse validation and 500 the whole tasks list.
    required = int((task.queue.required_annotators if task.queue else None) or 3)
    return {
        "dataset": task.dataset or "",
        "input_preview": input_preview(task.input_text),
        "submitted_count": int(task.submitted_count or 0),
        "required_annotators": required,
        "finalized_by_name": task.finalized_by_name or None,
        "finalized_at": format_iso(task.finalized_at),
        "id": str(task.id),
        "queue_id": str(task.queue_id),
        "task_name": (task.queue.task_name if task.queue else None) or "Task",
        "file_name": task.file_name or None,
        "batch_name": task.batch_name or "",
        "queue_name": (task.queue.name if task.queue else None) or "",
        "environment": "qa" if task.qa_queue_id else (task.environment or "production"),
        "status": task.status or "pending",
        "aht": format_aht(task.elapsed_seconds),
        "elapsed_seconds": task.elapsed_seconds or 0,
        "created_at": format_iso(task.created_at),
        "started_at": format_iso(task.started_at),
        "submitted_at": format_iso(task.submitted_at),
        "submitted_by": task.submitted_by or "",
    }


class TaskService:
    @staticmethod
    def list_tasks(
        db: Session,
        *,
        search: str | None = None,
        status: str | None = None,
        environment: str | None = None,
        queue: str | None = None,
        sort_by: str | None = None,
        sort_order: str | None = None,
        page: int | None = None,
        page_size: int | None = None,
    ):
        items, total, eff_page, eff_size = TaskRepository.get_tasks(
            db,
            search=search,
            status=status,
            environment=environment,
            queue=queue,
            sort_by=sort_by,
            sort_order=sort_order,
            page=page,
            page_size=page_size,
        )
        return [serialize_task(t) for t in items], total, eff_page, eff_size

    @staticmethod
    def get_queue_names(db: Session) -> list[str]:
        return TaskRepository.get_queue_names(db)

    @staticmethod
    def reinject_task(db: Session, task_id: str) -> bool:
        """Reopen an approved (finalised) task for QA review (SPEC2 §5.3).

        approved → "submitted": the task is routed back to the production
        queue's linked QA queue, every ``final_*`` field is cleared and both
        queues go back to "active" if they had been completed."""
        task = TaskRepository.get_task_by_id(db, task_id)
        if not task:
            return False

        # Only approved tasks can be reinjected.
        if task.status != "approved":
            return False

        queue = db.query(Queue).filter(Queue.id == task.queue_id).first()
        if not queue:
            return False

        qa_queue: Queue | None = None
        if queue.linked_qa_queue_id:
            qa_queue = db.query(Queue).filter(Queue.id == queue.linked_qa_queue_id).first()
        if qa_queue is None:
            qa_queue = (
                db.query(Queue)
                .filter(Queue.source_production_queue_id == queue.id, Queue.annotation_type == "qa")
                .first()
            )
        if qa_queue is None and task.qa_queue_id:
            qa_queue = db.query(Queue).filter(Queue.id == task.qa_queue_id).first()

        now = datetime.now(UTC)
        task.status = "submitted"
        task.qa_queue_id = qa_queue.id if qa_queue is not None else task.qa_queue_id
        task.final_data = None
        task.final_record = None
        task.finalized_by = None
        task.finalized_by_name = None
        task.finalized_at = None
        task.updated_at = now

        if queue.status == "completed":
            queue.status = "active"
            queue.updated_at = now
        if qa_queue is not None and qa_queue.status == "completed":
            qa_queue.status = "active"
            qa_queue.updated_at = now

        db.commit()
        return True

    @staticmethod
    def export_tasks_to_excel(
        db: Session,
        *,
        search: str | None = None,
        status: str | None = None,
        environment: str | None = None,
        queue: str | None = None,
        sort_by: str | None = None,
        sort_order: str | None = None,
    ) -> bytes:
        # Retrieve all filtered items without pagination
        items, _, _, _ = TaskRepository.get_tasks(
            db,
            search=search,
            status=status,
            environment=environment,
            queue=queue,
            sort_by=sort_by,
            sort_order=sort_order,
            page=None,
            page_size=None,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks Report"

        headers = [
            "S.No",
            "Dataset",
            "Prompt",
            "Task Name",
            "Batch Name",
            "Queue",
            "Type",
            "Status",
            "Submissions",
            "Finalized By",
            "Finalized At",
            "Created At",
            "Started At",
            "Submitted At",
            "Submitted By",
            "Elapsed",
            "AHT",
            "Timer",
            "Declined Reason",
        ]

        # Styling
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="18396B", end_color="18396B", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for idx, t in enumerate(items, 1):
            required = int((t.queue.required_annotators if t.queue else None) or 3)
            row_data = [
                idx,
                t.dataset or "",
                input_preview(t.input_text, 200),
                t.file_name or "",
                t.batch_name or "",
                t.queue.name if t.queue else "",
                ("qa" if t.qa_queue_id else (t.environment or "production")).capitalize(),
                (t.status or "pending").capitalize(),
                f"{int(t.submitted_count or 0)}/{required}",
                t.finalized_by_name or "",
                format_excel_dt(t.finalized_at),
                format_excel_dt(t.created_at),
                format_excel_dt(t.started_at),
                format_excel_dt(t.submitted_at),
                t.submitted_by or "",
                t.elapsed_seconds or 0,
                format_aht_hhmmss(t.elapsed_seconds),
                t.timer_seconds or 0,
                t.declined_reason or "",
            ]
            ws.append(row_data)

        # Autofit columns (the Prompt column is capped so a 200-char preview
        # doesn't produce an unusable sheet).
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

        # Write to memory stream
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
