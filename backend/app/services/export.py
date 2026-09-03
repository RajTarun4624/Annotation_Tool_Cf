"""Queue result exports — JSONL / JSON / Excel (SPEC2 §5.2).

Every export is computed over the PRODUCTION queue (a QA queue id resolves to
its source production queue). Scope ``final`` = approved tasks and their
stored ``final_record``; scope ``all`` = every task in sequence order, with
non-approved tasks rendered through ``build_record`` over whatever has been
submitted so far, plus a ``status`` key.
"""

from __future__ import annotations

import io
import json
import os
import re
import tempfile
import uuid
from collections.abc import Iterator
from datetime import datetime
from typing import Any

from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, selectinload
from starlette.background import BackgroundTask

from app.core.database import SessionLocal
from app.models import TaskAnnotation
from app.models.queue import Queue
from app.models.task import Task
from app.services.consensus import build_record, compute_consensus, normalise_annotation

EXPORT_SCOPES = ("final", "all")
EXPORT_PAGE_SIZE = 500

META_COLUMNS = (
    "data_type",
    "data_length_chars",
    "data_length_bucket",
    "data_structure",
    "attack_type",
    "attack_subcategory",
    "domain",
    "role",
    "verified",
    "language",
    "source_description",
    "document_edited",
)
ANNOTATOR_COLUMNS = (
    "name",
    "attack_type",
    "attack_subcategory",
    "jailbreak",
    "prompt_injection",
    "prompt_leakage",
    "severity_J",
    "severity_I",
    "severity_L",
    "intention",
    "verified",
)
AGREEMENT_COLUMNS = (
    "attack_type",
    "attack_subcategory",
    "output_jailbreak",
    "output_prompt_injection",
    "output_prompt_leakage",
    "severity_J",
    "severity_I",
    "severity_L",
    "intention",
    "verified",
)
TAIL_COLUMNS = (
    "output_jailbreak",
    "output_prompt_injection",
    "output_prompt_leakage",
    "severity_J",
    "severity_I",
    "severity_L",
    "intention",
    "source",
    "status",
    "finalized_by",
    "finalized_at",
    "qa_notes",
    "json",
)


def slugify(name: str | None) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")
    return slug or "queue"


def export_filename(queue: Queue, fmt: str) -> str:
    return f"{slugify(queue.name)}_results.{fmt}"


def _parse_uuid(value: Any) -> uuid.UUID | None:
    if isinstance(value, uuid.UUID):
        return value
    try:
        return uuid.UUID(str(value))
    except (ValueError, TypeError, AttributeError):
        return None


def resolve_production_queue(db: Session, queue_id: str) -> Queue | None:
    """The production queue behind ``queue_id`` (itself, or the source of a
    QA queue). None when the id is unknown."""
    queue_uuid = _parse_uuid(queue_id)
    if queue_uuid is None:
        return None
    queue = db.query(Queue).filter(Queue.id == queue_uuid).first()
    if queue is None:
        return None
    if (queue.annotation_type or "production") == "qa":
        source = None
        if queue.source_production_queue_id:
            source = db.query(Queue).filter(Queue.id == queue.source_production_queue_id).first()
        if source is None:
            source = db.query(Queue).filter(Queue.linked_qa_queue_id == queue.id).first()
        return source
    return queue


def _submitted_annotations(task: Task) -> list[TaskAnnotation]:
    subs = [a for a in (task.annotations or []) if (a.status or "") == "submitted"]
    subs.sort(key=lambda a: (a.submitted_at is None, a.submitted_at or datetime.min, a.created_at or datetime.min))
    return subs


def _empty_final() -> dict[str, Any]:
    try:
        return normalise_annotation({})
    except Exception:  # noqa: BLE001 - never let an empty task break the export
        return {}


def _reorder_like(stored: Any, template: Any) -> Any:
    """Return ``stored`` with dict keys ordered like ``template`` (recursively);
    keys the template lacks keep their stored order at the end. Values always
    come from ``stored`` — JSONB drops key order, this puts it back."""
    if not isinstance(stored, dict) or not isinstance(template, dict):
        return stored
    ordered: dict[str, Any] = {}
    for key, tmpl_value in template.items():
        if key in stored:
            ordered[key] = _reorder_like(stored[key], tmpl_value)
    for key, value in stored.items():
        if key not in ordered:
            ordered[key] = value
    return ordered


def _stored_record(task: Task, annotations: list[TaskAnnotation]) -> dict[str, Any]:
    """The approved task's ``final_record`` in the customer's key order."""
    stored = dict(task.final_record)
    try:
        template = build_record(task, annotations, task.final_data or {})
    except Exception:  # noqa: BLE001 - ordering is cosmetic, never lose the record
        return stored
    return _reorder_like(stored, template)


def _record_for(task: Task, annotations: list[TaskAnnotation]) -> dict[str, Any]:
    """Record for a task that has not been finalised: the majority (when at
    least one annotation exists) stands in for the final annotation."""
    if annotations:
        consensus = compute_consensus(task, annotations)
        final = consensus.get("majority") or _empty_final()
    else:
        final = task.final_data or _empty_final()
    return build_record(task, annotations, final)


def _row_for(task: Task) -> dict[str, Any]:
    annotations = _submitted_annotations(task)
    if task.status == "approved" and task.final_record:
        record = _stored_record(task, annotations)
    elif task.status == "approved" and task.final_data:
        record = build_record(task, annotations, task.final_data)
    else:
        record = _record_for(task, annotations)
    return {"task": task, "record": record, "annotations": annotations}


def iter_export_rows(db: Session, queue: Queue, scope: str = "final", page_size: int = EXPORT_PAGE_SIZE) -> Iterator[dict[str, Any]]:
    """Yield {"task", "record", "annotations"} for the production ``queue`` in
    sequence order, reading ``page_size`` tasks at a time so memory stays flat
    however many tasks the queue holds."""
    scope = scope if scope in EXPORT_SCOPES else "final"
    base = (
        db.query(Task)
        .options(selectinload(Task.annotations))
        .filter(Task.queue_id == queue.id)
    )
    if scope == "final":
        base = base.filter(Task.status == "approved")
    base = base.order_by(Task.sequence.asc(), Task.created_at.asc(), Task.id.asc())
    offset = 0
    while True:
        page = base.offset(offset).limit(page_size).all()
        if not page:
            return
        for task in page:
            yield _row_for(task)
        if len(page) < page_size:
            return
        offset += page_size
        db.expire_all()  # let the previous page be garbage-collected


def collect_export_rows(db: Session, queue: Queue, scope: str = "final") -> list[dict[str, Any]]:
    """[{"task", "record", "annotations"}] for the production ``queue``
    (materialised; prefer ``iter_export_rows`` for large queues)."""
    return list(iter_export_rows(db, queue, scope))


def records_only(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row["record"] for row in rows]


def _json_default(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, uuid.UUID):
        return str(value)
    return str(value)


def to_jsonl(records: list[dict[str, Any]]) -> bytes:
    lines = [json.dumps(r, ensure_ascii=False, default=_json_default) for r in records]
    return ("\n".join(lines) + ("\n" if lines else "")).encode("utf-8")


def to_json(records: list[dict[str, Any]]) -> bytes:
    return json.dumps(records, ensure_ascii=False, indent=2, default=_json_default).encode("utf-8")


def xlsx_headers(required_annotators: int) -> list[str]:
    headers: list[str] = ["dataset", "input", *META_COLUMNS]
    for k in range(1, max(1, int(required_annotators or 1)) + 1):
        headers.extend(f"annotator_{k}_{col}" for col in ANNOTATOR_COLUMNS)
    headers.extend(f"agreement_{col}" for col in AGREEMENT_COLUMNS)
    headers.append("consensus_reached")
    headers.extend(TAIL_COLUMNS)
    return headers


def _cell(value: Any) -> Any:
    """Excel-safe cell value: scalars pass through, structures are JSON."""
    if value is None:
        return ""
    if isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, (list, tuple)):
        return ", ".join(str(v) for v in value)
    return json.dumps(value, ensure_ascii=False, default=_json_default)


def _xlsx_row(row: dict[str, Any], required_annotators: int) -> list[Any]:
    task: Task = row["task"]
    record: dict[str, Any] = row["record"]
    annotations: list[TaskAnnotation] = row["annotations"]

    meta = record.get("meta_data") or {}
    output = record.get("output") or {}
    annotation = record.get("annotation") or {}
    severity = annotation.get("severity") or {}
    agreement = record.get("inter_annotator_agreement") or {}

    values: list[Any] = [record.get("dataset") or task.dataset or "", record.get("input") or task.input_text or ""]
    values.extend(_cell(meta.get(col)) for col in META_COLUMNS)

    for k in range(1, max(1, int(required_annotators or 1)) + 1):
        block = record.get(f"annotator_{k}") or {}
        ann = annotations[k - 1] if k - 1 < len(annotations) else None
        block_output = block.get("output") or {}
        block_severity = block.get("severity") or {}
        values.extend([
            _cell(ann.user_name if ann is not None else ""),
            _cell(block.get("attack_type")),
            _cell(block.get("attack_subcategory")),
            _cell(block_output.get("jailbreak")),
            _cell(block_output.get("prompt_injection")),
            _cell(block_output.get("prompt_leakage")),
            _cell(block_severity.get("J")),
            _cell(block_severity.get("I")),
            _cell(block_severity.get("L")),
            _cell(block.get("intention")),
            _cell(block.get("verified")),
        ])

    values.extend(_cell(agreement.get(col)) for col in AGREEMENT_COLUMNS)
    values.append(_cell(agreement.get("consensus_reached")))

    values.extend([
        _cell(output.get("jailbreak")),
        _cell(output.get("prompt_injection")),
        _cell(output.get("prompt_leakage")),
        _cell(severity.get("J")),
        _cell(severity.get("I")),
        _cell(severity.get("L")),
        _cell(annotation.get("intention")),
        _cell(record.get("source")),
        task.status or "pending",
        task.finalized_by_name or "",
        _cell(task.finalized_at),
        task.qa_notes or "",
        json.dumps(record, ensure_ascii=False, default=_json_default),
    ])
    return values


def to_xlsx(rows: list[dict[str, Any]], required_annotators: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = xlsx_headers(required_annotators)
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append(_xlsx_row(row, required_annotators))

    for col, header in enumerate(headers, start=1):
        if header in ("input", "json"):
            width = 60
        elif header in ("dataset", "source_description", "qa_notes"):
            width = 28
        else:
            width = max(12, min(len(header) + 2, 30))
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_queue(db: Session, queue: Queue, fmt: str, scope: str = "final") -> tuple[bytes, str, str]:
    """Return (body, media_type, filename) for ``fmt`` in jsonl | json | xlsx
    (materialised in memory; the API uses ``stream_export_response``)."""
    rows = collect_export_rows(db, queue, scope)
    if fmt == "jsonl":
        return to_jsonl(records_only(rows)), "application/x-ndjson", export_filename(queue, "jsonl")
    if fmt == "json":
        return to_json(records_only(rows)), "application/json", export_filename(queue, "json")
    if fmt == "xlsx":
        body = to_xlsx(rows, int(queue.required_annotators or 3))
        return (
            body,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            export_filename(queue, "xlsx"),
        )
    raise ValueError(f"Unsupported export format: {fmt}")


# ─── Streaming exports (constant memory, download starts immediately) ──────

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _stream_jsonl(queue_id: uuid.UUID, scope: str) -> Iterator[bytes]:
    db = SessionLocal()
    try:
        queue = db.query(Queue).filter(Queue.id == queue_id).first()
        if queue is None:
            return
        for row in iter_export_rows(db, queue, scope):
            yield (json.dumps(row["record"], ensure_ascii=False, default=_json_default) + "\n").encode("utf-8")
    finally:
        db.close()


def _stream_json(queue_id: uuid.UUID, scope: str) -> Iterator[bytes]:
    db = SessionLocal()
    try:
        queue = db.query(Queue).filter(Queue.id == queue_id).first()
        yield b"[\n"
        if queue is not None:
            first = True
            for row in iter_export_rows(db, queue, scope):
                chunk = json.dumps(row["record"], ensure_ascii=False, indent=2, default=_json_default)
                yield (("" if first else ",\n") + chunk).encode("utf-8")
                first = False
        yield b"\n]\n"
    finally:
        db.close()


def write_xlsx_file(db: Session, queue: Queue, scope: str, path: str) -> None:
    """Write the workbook row by row (openpyxl write-only mode) to ``path``."""
    required = int(queue.required_annotators or 3)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Results")
    headers = xlsx_headers(required)
    for col, header in enumerate(headers, start=1):
        if header in ("input", "json"):
            width = 60
        elif header in ("dataset", "source_description", "qa_notes"):
            width = 28
        else:
            width = max(12, min(len(header) + 2, 30))
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_cells = []
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        header_cells.append(cell)
    ws.append(header_cells)
    for row in iter_export_rows(db, queue, scope):
        ws.append(_xlsx_row(row, required))
    wb.save(path)


def stream_export_response(queue: Queue, fmt: str, scope: str = "final"):
    """FastAPI response for a queue export: JSON/JSONL are streamed straight
    from the database in pages; XLSX is written to a temp file in write-only
    mode and served with FileResponse (deleted afterwards)."""
    scope = scope if scope in EXPORT_SCOPES else "final"
    filename = export_filename(queue, fmt)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    if fmt == "jsonl":
        return StreamingResponse(_stream_jsonl(queue.id, scope), media_type="application/x-ndjson", headers=headers)
    if fmt == "json":
        return StreamingResponse(_stream_json(queue.id, scope), media_type="application/json", headers=headers)
    if fmt == "xlsx":
        db = SessionLocal()
        tmp = tempfile.NamedTemporaryFile(prefix="export_", suffix=".xlsx", delete=False)
        tmp.close()
        try:
            fresh = db.query(Queue).filter(Queue.id == queue.id).first() or queue
            write_xlsx_file(db, fresh, scope, tmp.name)
        except Exception:
            os.unlink(tmp.name)
            raise
        finally:
            db.close()
        return FileResponse(
            tmp.name,
            media_type=XLSX_MIME,
            headers=headers,
            filename=filename,
            background=BackgroundTask(_unlink_quiet, tmp.name),
        )
    raise ValueError(f"Unsupported export format: {fmt}")


def _unlink_quiet(path: str) -> None:
    try:
        os.unlink(path)
    except OSError:
        pass


def export_single_task(db: Session, task: Task, fmt: str = "json") -> tuple[bytes, str, str]:
    """Return (body, media_type, filename) for a single task export."""
    annotations = _submitted_annotations(task)
    if task.status == "approved" and task.final_record:
        record = _stored_record(task, annotations)
    elif task.status == "approved" and task.final_data:
        record = build_record(task, annotations, task.final_data)
    else:
        record = _record_for(task, annotations)

    row = {"task": task, "record": record, "annotations": annotations}
    prod = task.queue
    if prod is None and task.queue_id:
        prod = db.query(Queue).filter(Queue.id == task.queue_id).first()
    required = int(prod.required_annotators or 3) if prod else 3

    task_slug = slugify(task.dataset or str(task.id)[:8])
    filename = f"task_{task_slug}.{fmt}"

    if fmt == "jsonl":
        return to_jsonl([record]), "application/x-ndjson", filename
    if fmt == "json":
        return to_json([record]), "application/json", filename
    if fmt == "xlsx":
        body = to_xlsx([row], required)
        return (
            body,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename,
        )
    raise ValueError(f"Unsupported export format: {fmt}")

