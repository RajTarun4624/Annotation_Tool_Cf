"""Spreadsheet (xlsx / csv) → prompt rows for queue creation (SPEC2 §5.1).

``parse_sheet`` never touches the database: it turns the uploaded bytes into
the ``{"columns", "detected", "total_rows", "skipped_rows", "rows"}`` payload
the Create Queue modal previews and then sends back verbatim as
``QueueCreateRequest.tasks``. Every failure is an ``HTTPException(400)`` so
the client can show the message as-is.
"""

from __future__ import annotations

import csv
import io
import os
import re
from datetime import date, datetime
from typing import Any

from fastapi import HTTPException, status
from openpyxl import load_workbook

MAX_SHEET_BYTES = 20 * 1024 * 1024  # 20 MB
MAX_PROMPT_ROWS = 5000

INPUT_HEADERS = ("input", "prompt", "text", "prompt_text", "input_text", "content", "message")
DATASET_HEADERS = ("dataset", "data_set", "task_id", "taskid", "id", "name")
META_HEADERS = (
    "data_type",
    "data_structure",
    "attack_type",
    "attack_subcategory",
    "domain",
    "role",
    "language",
    "source",
    "source_description",
    "verified",
    "document_edited",
    "intention",
)
BOOL_HEADERS = {"verified", "document_edited"}
_TRUE_WORDS = {"true", "yes", "y", "1", "verified", "edited"}
_FALSE_WORDS = {"false", "no", "n", "0", "unverified", ""}


def _bad_request(detail: str) -> HTTPException:
    return HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=detail)


def normalise_header(raw: Any) -> str:
    """lowercase, trim, spaces/hyphens → underscore ("Prompt Text" → "prompt_text")."""
    text = "" if raw is None else str(raw)
    text = text.replace("﻿", "").strip().lower()
    text = re.sub(r"[\s\-]+", "_", text)
    return text


def sanitise_prefix(name_prefix: str | None) -> str:
    """Dataset-id prefix restricted to [a-z0-9_]+ (default "task")."""
    text = (name_prefix or "").strip().lower()
    text = re.sub(r"[^a-z0-9_]+", "_", text).strip("_")
    text = re.sub(r"_+", "_", text)
    return text or "task"


def _cell_to_text(value: Any) -> str:
    """Stringify a spreadsheet cell without the float noise openpyxl adds
    ("12.0" → "12") and without Python's datetime repr."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value)
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _cell_to_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = _cell_to_text(value).lower()
    if text in _TRUE_WORDS:
        return True
    if text in _FALSE_WORDS:
        return False
    return None


def _read_xlsx(file_bytes: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises many unrelated types
        raise _bad_request("Could not read the workbook. Is it a valid .xlsx file?") from exc
    try:
        sheet = workbook.worksheets[0] if workbook.worksheets else workbook.active
        if sheet is None:
            raise _bad_request("The workbook has no worksheets.")
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _read_csv(file_bytes: bytes) -> list[list[Any]]:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = file_bytes.decode("cp1252")
        except UnicodeDecodeError as exc:
            raise _bad_request("Could not decode the CSV file (expected UTF-8).") from exc
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    return [list(row) for row in reader]


def _is_blank_row(row: list[Any]) -> bool:
    return all(_cell_to_text(cell) == "" for cell in row)


def parse_sheet(file_bytes: bytes, filename: str, name_prefix: str | None = None) -> dict[str, Any]:
    """Parse an .xlsx / .csv upload into prompt rows.

    Returns ``{"columns": [...original headers...],
               "detected": {"input": header, "dataset": header | None},
               "total_rows": n_kept, "skipped_rows": n_skipped,
               "rows": [{"dataset", "input", "meta_data"}, ...]}``.
    """
    if not file_bytes:
        raise _bad_request("The uploaded file is empty.")
    if len(file_bytes) > MAX_SHEET_BYTES:
        raise _bad_request("The spreadsheet is larger than 20 MB.")

    ext = os.path.splitext(filename or "")[1].lower()
    if ext == ".xlsx":
        raw_rows = _read_xlsx(file_bytes)
    elif ext == ".csv":
        raw_rows = _read_csv(file_bytes)
    else:
        raise _bad_request("Unsupported file type. Upload a .xlsx or .csv spreadsheet.")

    # First non-blank row is the header row.
    header_index = next((i for i, row in enumerate(raw_rows) if not _is_blank_row(row)), None)
    if header_index is None:
        raise _bad_request("The spreadsheet has no header row.")

    original_headers = [_cell_to_text(cell) for cell in raw_rows[header_index]]
    # Trim trailing empty header cells (openpyxl pads to the widest row).
    while original_headers and original_headers[-1] == "":
        original_headers.pop()
    normalised = [normalise_header(h) for h in original_headers]

    def find_column(candidates: tuple[str, ...]) -> int | None:
        for candidate in candidates:
            if candidate in normalised:
                return normalised.index(candidate)
        return None

    input_index = find_column(INPUT_HEADERS)
    if input_index is None:
        found = ", ".join(h for h in original_headers if h) or "(none)"
        raise _bad_request(f"Could not find an input/prompt column. Found headers: {found}")
    dataset_index = find_column(DATASET_HEADERS)

    meta_columns: list[tuple[str, int]] = []
    for key in META_HEADERS:
        if key in normalised:
            idx = normalised.index(key)
            if idx not in (input_index, dataset_index):
                meta_columns.append((key, idx))

    prefix = sanitise_prefix(name_prefix)
    rows: list[dict[str, Any]] = []
    skipped = 0

    for raw in raw_rows[header_index + 1:]:
        if _is_blank_row(raw):
            continue
        cells = list(raw) + [None] * (len(original_headers) - len(raw))

        input_text = _cell_to_text(cells[input_index]).strip()
        if not input_text:
            skipped += 1
            continue

        if len(rows) >= MAX_PROMPT_ROWS:
            raise _bad_request("Sheet has more than 5000 prompt rows.")

        meta: dict[str, Any] = {}
        for key, idx in meta_columns:
            value = cells[idx]
            if key in BOOL_HEADERS:
                coerced = _cell_to_bool(value)
                if coerced is not None and _cell_to_text(value) != "":
                    meta[key] = coerced
                continue
            text = _cell_to_text(value).strip()
            if text:
                meta[key] = text

        dataset = _cell_to_text(cells[dataset_index]).strip() if dataset_index is not None else ""
        if not dataset:
            dataset = f"{prefix}_{len(rows) + 1:04d}"

        rows.append({"dataset": dataset, "input": input_text, "meta_data": meta})

    return {
        "columns": original_headers,
        "detected": {
            "input": original_headers[input_index],
            "dataset": original_headers[dataset_index] if dataset_index is not None else None,
        },
        "total_rows": len(rows),
        "skipped_rows": skipped,
        "rows": rows,
    }
